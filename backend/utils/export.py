import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import mm
from datetime import datetime

def generate_excel(report_type: str, data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace("-", " ").title()
    
    # Header style
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{report_type.replace('-', ' ').upper()} REPORT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)
    
    row = 4
    
    if report_type == "pnl":
        headers = ["Metric", "Amount"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        row += 1
        for metric, value in [
            ("Total Revenue", data.get("total_revenue", 0)),
            ("Total COGS", data.get("total_cogs", 0)),
            ("Gross Profit", data.get("gross_profit", 0)),
            ("Total Expenses", data.get("total_expenses", 0)),
            ("Net Profit", data.get("net_profit", 0)),
            ("Margin %", data.get("margin_percentage", 0)),
        ]:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        if data.get("revenue_by_outlet"):
            row += 1
            ws.cell(row=row, column=1, value="Revenue by Outlet")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            for item in data["revenue_by_outlet"]:
                ws.cell(row=row, column=1, value=item.get("outlet_name", ""))
                ws.cell(row=row, column=2, value=item.get("revenue", 0))
                row += 1
    
    elif report_type == "cashflow":
        headers = ["Date", "Inflow", "Outflow", "Net"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        for item in data.get("daily_cashflow", []):
            ws.cell(row=row, column=1, value=item.get("date", ""))
            ws.cell(row=row, column=2, value=item.get("inflow", 0))
            ws.cell(row=row, column=3, value=item.get("outflow", 0))
            ws.cell(row=row, column=4, value=item.get("net", 0))
            row += 1
    
    elif report_type == "balance-sheet":
        ws.cell(row=row, column=1, value="ASSETS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        assets = data.get("assets", {})
        for key, val in assets.items():
            ws.cell(row=row, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row, column=2, value=val)
            row += 1
    
    elif report_type == "inventory-valuation":
        headers = ["Item", "Category", "UOM", "Quantity", "Cost/Unit", "Value", "Outlet"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        for item in data.get("items", []):
            ws.cell(row=row, column=1, value=item.get("item_name", ""))
            ws.cell(row=row, column=2, value=item.get("category", ""))
            ws.cell(row=row, column=3, value=item.get("uom", ""))
            ws.cell(row=row, column=4, value=item.get("quantity", 0))
            ws.cell(row=row, column=5, value=item.get("cost_per_unit", 0))
            ws.cell(row=row, column=6, value=item.get("value", 0))
            ws.cell(row=row, column=7, value=item.get("outlet_name", ""))
            row += 1
        row += 1
        ws.cell(row=row, column=5, value="Total:")
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6, value=data.get("total_value", 0))
    
    elif report_type == "inventory-movements":
        headers = ["Type", "Count", "Total Quantity"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row += 1
        for item in data.get("summary", []):
            ws.cell(row=row, column=1, value=item.get("type", ""))
            ws.cell(row=row, column=2, value=item.get("count", 0))
            ws.cell(row=row, column=3, value=item.get("total_quantity", 0))
            row += 1
    
    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def generate_pdf(report_type: str, data: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = []
    
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=12)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.grey)
    
    elements.append(Paragraph(f"{report_type.replace('-', ' ').upper()} REPORT", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 12))
    
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')]),
    ])
    
    if report_type == "pnl":
        table_data = [["Metric", "Amount"]]
        for metric, value in [
            ("Total Revenue", f"{data.get('total_revenue', 0):,.0f}"),
            ("Total COGS", f"{data.get('total_cogs', 0):,.0f}"),
            ("Gross Profit", f"{data.get('gross_profit', 0):,.0f}"),
            ("Total Expenses", f"{data.get('total_expenses', 0):,.0f}"),
            ("Net Profit", f"{data.get('net_profit', 0):,.0f}"),
            ("Margin %", f"{data.get('margin_percentage', 0):.1f}%"),
        ]:
            table_data.append([metric, value])
        t = Table(table_data, colWidths=[200, 150])
        t.setStyle(table_style)
        elements.append(t)
    
    elif report_type == "inventory-valuation":
        table_data = [["Item", "Category", "Qty", "Cost/Unit", "Value"]]
        for item in data.get("items", [])[:50]:  # Limit to 50 for PDF
            table_data.append([
                item.get("item_name", "")[:25],
                item.get("category", ""),
                f"{item.get('quantity', 0):,.1f}",
                f"{item.get('cost_per_unit', 0):,.0f}",
                f"{item.get('value', 0):,.0f}",
            ])
        table_data.append(["", "", "", "Total:", f"{data.get('total_value', 0):,.0f}"])
        t = Table(table_data, colWidths=[120, 80, 60, 80, 80])
        t.setStyle(table_style)
        elements.append(t)
    
    else:
        elements.append(Paragraph(f"Report data for: {report_type}", styles['Normal']))
    
    doc.build(elements)
    return buffer.getvalue()
